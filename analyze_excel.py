"""Script to analyze the new UNIFORMES.xlsx structure."""
import pandas as pd
import openpyxl

file_path = 'sources/UNIFORMES.xlsx'
print('ANALYZING:', file_path)
print('=' * 80)

# Load with openpyxl for detailed structure
wb = openpyxl.load_workbook(file_path, data_only=True)
print(f'\n📋 WORKSHEETS ({len(wb.sheetnames)}):')
for i, sheet_name in enumerate(wb.sheetnames, 1):
    ws = wb[sheet_name]
    print(f'  {i}. "{sheet_name}" - {ws.max_row} rows x {ws.max_column} cols')

# Analyze first sheet in detail
ws = wb.active
print(f'\n📊 Active sheet: "{ws.title}"')
print(f'Dimensions: {ws.dimensions}')

# Show first 15 rows with all columns
print('\n' + '=' * 80)
print('FIRST 15 ROWS (columns A-T):')
print('=' * 80)
for row_idx in range(1, min(16, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(21, ws.max_column + 1)):
        cell = ws.cell(row=row_idx, column=col_idx)
        val = cell.value
        if val is not None:
            val_str = str(val)[:20]
        else:
            val_str = ''
        row_data.append(val_str)
    print(f'Row {row_idx:2d}: ' + ' | '.join(row_data))

# Column headers analysis
print('\n' + '=' * 80)
print('ALL COLUMN HEADERS:')
print('=' * 80)
for col_idx in range(1, ws.max_column + 1):
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    # Check rows 1-5 for header content
    for row in range(1, 6):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value:
            print(f'  {col_letter} (col {col_idx}, row {row}): {cell.value}')
            break

# Merged cells
print('\n📦 MERGED CELLS:')
if ws.merged_cells.ranges:
    for merged_range in list(ws.merged_cells.ranges)[:20]:
        print(f'  {merged_range}')
else:
    print('  None')

# Read with pandas for data analysis
print('\n' + '=' * 80)
print('PANDAS DATA VIEW:')
print('=' * 80)
df = pd.read_excel(file_path, header=None)
print(df.head(20).to_string())

print('\n' + '=' * 80)
print('DATA ROWS SAMPLE (rows 10-20):')
print('=' * 80)
if len(df) > 10:
    print(df.iloc[9:20].to_string())
